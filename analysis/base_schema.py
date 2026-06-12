import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker
import numpy as np
from collections import Counter

matplotlib.use("Agg")

wb = openpyxl.load_workbook("NIH IC Efforts Landscape 2026.04.22.xlsx")
ws = wb["Ecosystems"]
headers = [cell.value for cell in ws[1]]
n_platforms = ws.max_row - 1

col = headers.index("Base schema") + 1

raw_values = {}
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=2).value
    val = ws.cell(row=r, column=col).value
    raw_values[name] = val

# Normalize known aliases
aliases = {
    "DATaset Metadata Model (DATMM)": "DATMM",
    "DATMM": "DATMM",
}

def normalize(s):
    return aliases.get(s.strip().rstrip(","), s.strip().rstrip(","))

# Categorize each platform
custom, unknown, named, none_val = [], [], [], []
schema_list = []

for name, val in raw_values.items():
    if val is None:
        none_val.append(name)
    else:
        parts = [normalize(p) for p in str(val).split(",") if p.strip()]
        is_custom = any("custom" in p.lower() or "encode dcc" in p.lower() or "brics" in p.lower() for p in parts)
        is_unknown = all(p.lower() == "unknown" for p in parts)
        named_parts = [p for p in parts if p.lower() not in ("unknown",)
                       and "custom" not in p.lower()
                       and "encode dcc" not in p.lower()
                       and "brics" not in p.lower()]
        if is_unknown:
            unknown.append(name)
        elif is_custom and not named_parts:
            custom.append(name)
        elif is_custom and named_parts:
            custom.append(name)
            named.append(name)
            schema_list.extend(named_parts)
        else:
            named.append(name)
            schema_list.extend(named_parts)

counts = Counter(schema_list)

print(f"Total platforms: {n_platforms}")
print(f"Custom schema only: {len([n for n in custom if n not in named])}")
print(f"Custom + named schema: {len([n for n in custom if n in named])}")
print(f"Named schema only: {len([n for n in named if n not in custom])}")
print(f"Unknown: {len(unknown)}")
print(f"None listed: {len(none_val)}")
print(f"\nNamed schema counts:")
for s, c in counts.most_common():
    print(f"  {s}: {c}")

# --- Figure 1: Donut chart of schema categories ---
custom_only = [n for n in custom if n not in named]
custom_and_named = [n for n in custom if n in named]
named_only = [n for n in named if n not in custom]

category_counts = {
    "Custom only": len(custom_only),
    "Custom + named": len(custom_and_named),
    "Named schema": len(named_only),
    "Unknown": len(unknown),
}

fig, axes = plt.subplots(1, 2, figsize=(13, 6))

# Donut
colors_donut = ["#4C72B0", "#DD8452", "#55A868", "#C44E52"]
wedges, texts, autotexts = axes[0].pie(
    category_counts.values(),
    labels=None,
    autopct=lambda p: f"{p:.0f}%\n(n={int(round(p * n_platforms / 100))})",
    colors=colors_donut,
    startangle=90,
    wedgeprops=dict(width=0.5),
    pctdistance=0.75,
)
for at in autotexts:
    at.set_fontsize(9)
axes[0].legend(category_counts.keys(), loc="lower center", bbox_to_anchor=(0.5, -0.12),
               fontsize=9, ncol=2)
axes[0].set_title("Base Schema Categories", fontsize=11, fontweight="bold")

# Bar chart of named schemas
if counts:
    schemas = [s for s, c in counts.most_common()]
    vals = [c for s, c in counts.most_common()]
    y = np.arange(len(schemas))
    axes[1].barh(y, vals, color="#4C72B0", edgecolor="white", linewidth=0.5)
    for i, (v, s) in enumerate(zip(vals, schemas)):
        axes[1].text(v + 0.05, i, str(v), va="center", fontsize=9, color="#333333")
    axes[1].set_yticks(y)
    axes[1].set_yticklabels(schemas, fontsize=9)
    axes[1].invert_yaxis()
    axes[1].set_xlabel("Number of platforms", fontsize=10)
    axes[1].set_title("Named Base Schemas", fontsize=11, fontweight="bold")
    axes[1].xaxis.set_major_locator(matplotlib.ticker.MaxNLocator(integer=True))
    axes[1].set_xlim(0, max(vals) + 1)
    axes[1].spines["top"].set_visible(False)
    axes[1].spines["right"].set_visible(False)

fig.suptitle(f"Base Schema Usage Across NIH IC Data Ecosystem Platforms (n={n_platforms})",
             fontsize=12, fontweight="bold", y=1.01)
plt.tight_layout()
plt.savefig("analysis/base_schema.png", dpi=300, bbox_inches="tight")
plt.savefig("analysis/base_schema.pdf", bbox_inches="tight")
print("Saved: analysis/base_schema.png and .pdf")
