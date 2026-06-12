import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker
import numpy as np
from collections import Counter

matplotlib.use("Agg")

wb = openpyxl.load_workbook("NIH IC Efforts Landscape 2026.04.22.xlsx")
ws = wb["Ecosystems"]
headers = [cell.value for cell in ws[1]]
n_platforms = ws.max_row - 1

col = headers.index("Ontologies") + 1

# Normalize known aliases to a canonical name
aliases = {
    "Chemical Entities of Biological Interest (ChEBI)": "ChEBI",
    "chebi": "ChEBI",
    "Gene Ontology (GO)": "Gene Ontology (GO)",
    "GO (Gene Ontology)": "Gene Ontology (GO)",
    "go": "Gene Ontology (GO)",
    "Disease Ontology (DO)": "Disease Ontology (DO)",
    "Human Phenotype Ontology (HPO)": "Human Phenotype Ontology (HPO)",
    "Human Phenotype Ontology (HP)": "Human Phenotype Ontology (HPO)",
    "MONDO Disease Ontology (MONDO)": "MONDO Disease Ontology",
    "Monarch Disease Ontology (Mondo)": "MONDO Disease Ontology",
    "Ontology for Biomedical Investigations (OBI)": "Ontology for Biomedical Investigations (OBI)",
    "Uber Anatomy Ontology (UBERON)": "UBERON",
    "UBERON (Uber Anatomy Ontology)": "UBERON",
    "Medical Subject Headings (MeSH)": "MeSH",
    "Systematized Nomenclature of Medicine Clinical Terms (SNOMED-CT)": "SNOMED-CT",
    "Systematized Nomenclature of Medicine Clinical Terms (SNOMED CT)": "SNOMED-CT",
    "Logical Observation Identifiers Names and Codes (LOINC)": "LOINC",
    "NCBI Taxonomy": "NCBI Taxonomy",
    "NCBI Taxon": "NCBI Taxonomy",
    "Drug Target Ontology (DTO)": "Drug Target Ontology (DTO)",
    "Cell Line Ontology (CLO)": "Cell Line Ontology (CLO)",
    "Data Use Ontology (DUO)": "Data Use Ontology (DUO)",
    "NIH Common Data Elements": "NIH Common Data Elements",
    "National Institutes of Health Common Data Elements (NIH CDEs)": "NIH Common Data Elements",
    "NIH CDEs": "NIH Common Data Elements",
    "Human Phenotype Ontology (HPO)": "Human Phenotype Ontology (HPO)",
    "Foundational Model of Anatomy (FMA)": "Foundational Model of Anatomy (FMA)",
    "FMA (Foundational Model of Anatomy)": "Foundational Model of Anatomy (FMA)",
    "PubChem": "PubChem",
    "Observational Medical Outcomes Partnership (OMOP)": "OMOP",
    "Observational Health Data Sciences and Informatics (OHDSI) OMOP Common Data Model": "OMOP",
    "UniProt": "UniProt",
    "Universal Protein Knowledgebase (UniProtKB)": "UniProt",
    "NCBI Gene": "NCBI Gene",
    "National Center for Biotechnology Information Gene (NCBI Gene)": "NCBI Gene",
    "Neuroscience Information Framework": "NIF/NIFSTD",
    "Neuroscience Information Framework Standard Ontology (NIFSTD)": "NIF/NIFSTD",
    "NIF Standardized Ontologies": "NIF/NIFSTD",
}

def normalize(s):
    s = s.strip().rstrip(",").strip()
    return aliases.get(s, s)

none_platforms, ontology_list, platform_counts = [], [], {}

for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=2).value
    val = ws.cell(row=r, column=col).value
    if val is None or str(val).strip().lower() in ("none/unknown", "none", "unknown"):
        none_platforms.append(name)
        platform_counts[name] = []
    else:
        parts = [normalize(p) for p in str(val).split(",") if p.strip()]
        # Remove empty or whitespace-only after normalize
        parts = [p for p in parts if p]
        ontology_list.extend(parts)
        platform_counts[name] = parts

counts = Counter(ontology_list)
n_with_ontologies = n_platforms - len(none_platforms)
n_unique = len(counts)
per_platform = [len(v) for v in platform_counts.values() if v]

print(f"Platforms with ontologies listed: {n_with_ontologies}/{n_platforms}")
print(f"Platforms with None/Unknown: {len(none_platforms)}")
print(f"Total unique ontologies: {n_unique}")
print(f"Avg ontologies per platform (those with any): {np.mean(per_platform):.1f} (range {min(per_platform)}-{max(per_platform)})")
print(f"\nOntology counts (appearing in >=2 platforms):")
for ont, cnt in counts.most_common():
    if cnt >= 2:
        print(f"  {ont}: {cnt}")
print(f"\nOntologies in only 1 platform: {sum(1 for c in counts.values() if c == 1)}")

# Plot: ontologies appearing in >=2 platforms
filtered = [(ont, cnt) for ont, cnt in counts.most_common() if cnt >= 2]
labels = [f[0] for f in filtered]
values = [f[1] for f in filtered]

n_single = sum(1 for c in counts.values() if c == 1)

fig, ax = plt.subplots(figsize=(10, 6))
y = np.arange(len(labels))
bars = ax.barh(y, values, color="#4C72B0", edgecolor="white", linewidth=0.5)

for bar, cnt in zip(bars, values):
    ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
            str(cnt), va="center", ha="left", fontsize=9, color="#333333")

ax.set_yticks(y)
ax.set_yticklabels(labels, fontsize=9.5)
ax.invert_yaxis()
ax.set_xlabel("Number of platforms", fontsize=11)
ax.set_title(
    f"Ontology Prevalence Across NIH IC Data Ecosystem Platforms\n"
    f"(n={n_platforms} platforms; ontologies present in ≥2 platforms shown)",
    fontsize=12, fontweight="bold", pad=28
)
ax.annotate(
    f"Note: {len(none_platforms)} platforms have no listed ontologies. "
    f"{n_single} additional ontologies appear in only 1 platform.",
    xy=(0.5, 1.0), xycoords="axes fraction",
    ha="center", va="bottom", fontsize=8, color="gray"
)
ax.set_xlim(0, max(values) + 1.5)
ax.xaxis.set_major_locator(matplotlib.ticker.MaxNLocator(integer=True))
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)

n_single = sum(1 for c in counts.values() if c == 1)

plt.tight_layout()
plt.savefig("analysis/ontologies.png", dpi=300, bbox_inches="tight")
plt.savefig("analysis/ontologies.pdf", bbox_inches="tight")
print("Saved: analysis/ontologies.png and .pdf")
