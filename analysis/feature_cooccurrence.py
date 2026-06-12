import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

matplotlib.use("Agg")

wb = openpyxl.load_workbook("NIH IC Efforts Landscape 2026.04.22.xlsx")
ws = wb["Ecosystems"]
headers = [cell.value for cell in ws[1]]
n_platforms = ws.max_row - 1

bool_cols = [
    "API", "Search", "Public Uptime Checking",
    "Metadata/data standardization", "Metadata/data validation", "Metadata augmentation",
    "Workspace", "Repositories",
    "Datasets", "Samples", "Tools", "Publications", "Clinical Trials",
    "Protocols", "Images", "Multimedia", "Grants", "Models", "Websites",
    "Reports/Analyses", "Posters/Conference Contribution",
    "Book/Chapter/Collection", "Software Source Code", "Institution/Researcher Profiles",
]

short_labels = {
    "API": "API",
    "Search": "Search",
    "Public Uptime Checking": "Uptime Checking",
    "Metadata/data standardization": "Standardization",
    "Metadata/data validation": "Validation",
    "Metadata augmentation": "Augmentation",
    "Workspace": "Workspace",
    "Repositories": "Repositories",
    "Datasets": "Datasets",
    "Samples": "Samples",
    "Tools": "Tools",
    "Publications": "Publications",
    "Clinical Trials": "Clinical Trials",
    "Protocols": "Protocols",
    "Images": "Images",
    "Multimedia": "Multimedia",
    "Grants": "Grants",
    "Models": "Models",
    "Websites": "Websites",
    "Reports/Analyses": "Reports/Analyses",
    "Posters/Conference Contribution": "Posters",
    "Book/Chapter/Collection": "Books/Collections",
    "Software Source Code": "Software Source Code",
    "Institution/Researcher Profiles": "Inst./Researcher Profiles",
}

# Build binary matrix
col_idx = {h: headers.index(h) for h in bool_cols}
data = {}
for h in bool_cols:
    data[h] = [
        1 if ws.cell(row=r, column=col_idx[h] + 1).value is True else 0
        for r in range(2, ws.max_row + 1)
    ]

df = pd.DataFrame(data)

# Remove features with zero variance (all same value) — uninformative for co-occurrence
df = df.loc[:, df.std() > 0]
remaining_cols = list(df.columns)

# Compute co-occurrence as % of platforms where both features are TRUE
n = len(df)
cooc = pd.DataFrame(index=remaining_cols, columns=remaining_cols, dtype=float)
for a in remaining_cols:
    for b in remaining_cols:
        cooc.loc[a, b] = (df[a] & df[b]).sum() / n * 100

# Sort by total TRUE count (descending) for readability
order = df.sum().sort_values(ascending=False).index.tolist()
cooc = cooc.loc[order, order]
tick_labels = [short_labels[c] for c in order]

# Plot
fig, ax = plt.subplots(figsize=(13, 11))
im = ax.imshow(cooc.values, cmap="YlOrRd", vmin=0, vmax=100, aspect="auto")

cbar = fig.colorbar(im, ax=ax, fraction=0.03, pad=0.02)
cbar.set_label("% of platforms where both features are present", fontsize=10)

ax.set_xticks(range(len(order)))
ax.set_yticks(range(len(order)))
ax.set_xticklabels(tick_labels, rotation=45, ha="right", fontsize=8.5)
ax.set_yticklabels(tick_labels, fontsize=8.5)

# Annotate cells with values
for i in range(len(order)):
    for j in range(len(order)):
        val = cooc.values[i, j]
        color = "white" if val > 65 else "black"
        ax.text(j, i, f"{val:.0f}", ha="center", va="center", fontsize=6.5, color=color)

ax.set_title(
    f"Feature Co-occurrence Across NIH IC Data Ecosystem Platforms (n={n_platforms})\n"
    f"Values show % of platforms where both features are present",
    fontsize=12, fontweight="bold", pad=12
)

plt.tight_layout()
plt.savefig("analysis/feature_cooccurrence.png", dpi=300, bbox_inches="tight")
plt.savefig("analysis/feature_cooccurrence.pdf", bbox_inches="tight")
print("Saved: analysis/feature_cooccurrence.png and .pdf")
