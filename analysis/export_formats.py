import openpyxl
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
from collections import Counter

matplotlib.use("Agg")

wb = openpyxl.load_workbook("NIH IC Efforts Landscape 2026.04.22.xlsx")
ws = wb["Ecosystems"]
headers = [cell.value for cell in ws[1]]
n_platforms = ws.max_row - 1

col = headers.index("Export Formats") + 1

# Parse all formats
all_formats = []
platforms_with_formats = 0
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=col).value
    if val and str(val).strip().lower() != "none":
        formats = [f.strip() for f in str(val).split(",")]
        all_formats.extend(formats)
        platforms_with_formats += 1

counts = Counter(all_formats)
print(f"Platforms with export formats: {platforms_with_formats}/{n_platforms}")
print(f"Unique formats: {len(counts)}")
print("\nAll format counts:")
for fmt, cnt in counts.most_common():
    print(f"  {fmt}: {cnt}")

# Plot formats appearing in >=2 platforms
filtered = [(fmt, cnt) for fmt, cnt in counts.most_common() if cnt >= 2]
filtered_sorted = sorted(filtered, key=lambda x: x[1], reverse=True)
labels = [f[0] for f in filtered_sorted]
values = [f[1] for f in filtered_sorted]

fig, ax = plt.subplots(figsize=(10, 7))
y = np.arange(len(labels))
bars = ax.barh(y, values, color="#4C72B0", edgecolor="white", linewidth=0.5)

for bar, cnt in zip(bars, values):
    ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
            str(cnt), va="center", ha="left", fontsize=9, color="#333333")

ax.set_yticks(y)
ax.set_yticklabels(labels, fontsize=10)
ax.invert_yaxis()
ax.set_xlabel("Number of platforms", fontsize=11)
ax.set_title(
    f"Export Format Prevalence Across NIH IC Data Ecosystem Platforms\n"
    f"(n={n_platforms} platforms; formats present in ≥2 platforms shown)",
    fontsize=12, fontweight="bold", pad=12
)
ax.set_xlim(0, max(values) + 2)
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)

# Footnote
n_unique = len(counts)
n_shown = len(labels)
n_no_export = n_platforms - platforms_with_formats
ax.text(0, -1.8, f"Note: {n_no_export} platforms have no listed export formats. "
        f"{n_unique - n_shown} additional formats appear in only 1 platform.",
        fontsize=8, color="gray", transform=ax.get_yaxis_transform())

plt.tight_layout()
plt.savefig("analysis/export_formats.png", dpi=300, bbox_inches="tight")
plt.savefig("analysis/export_formats.pdf", bbox_inches="tight")
print("Saved: analysis/export_formats.png and .pdf")
