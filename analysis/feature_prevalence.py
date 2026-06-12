import openpyxl
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

# Load data
wb = openpyxl.load_workbook("NIH IC Efforts Landscape 2026.04.22.xlsx")
ws = wb["Ecosystems"]
headers = [cell.value for cell in ws[1]]
n_platforms = ws.max_row - 1  # 29

bool_cols = [
    "API", "Search", "Public Uptime Checking",
    "Metadata/data standardization", "Metadata/data validation", "Metadata augmentation",
    "Workspace", "Repositories",
    "Datasets", "Samples", "Tools", "Publications", "Clinical Trials",
    "Protocols", "Images", "Multimedia", "Grants", "Models", "Websites",
    "Reports/Analyses", "Posters/Conference Contribution",
    "Book/Chapter/Collection", "Software Source Code", "Institution/Researcher Profiles",
]

# Group labels for visual separation
groups = {
    "Technical Features": ["API", "Search", "Public Uptime Checking"],
    "Metadata": ["Metadata/data standardization", "Metadata/data validation", "Metadata augmentation"],
    "Infrastructure": ["Workspace", "Repositories"],
    "Resource Types": [
        "Datasets", "Samples", "Tools", "Publications", "Clinical Trials",
        "Protocols", "Images", "Multimedia", "Grants", "Models", "Websites",
        "Reports/Analyses", "Posters/Conference Contribution",
        "Book/Chapter/Collection", "Software Source Code", "Institution/Researcher Profiles",
    ],
}

# Shorter display labels
short_labels = {
    "API": "API",
    "Search": "Search",
    "Public Uptime Checking": "Uptime Checking",
    "Metadata/data standardization": "Standardization",
    "Metadata/data validation": "Validation",
    "Metadata augmentation": "Augmentation",
    "Workspace": "Workspace",
    "Repositories": "Repositories",
    "Datasets": "Datasets",
    "Samples": "Samples",
    "Tools": "Tools",
    "Publications": "Publications",
    "Clinical Trials": "Clinical Trials",
    "Protocols": "Protocols",
    "Images": "Images",
    "Multimedia": "Multimedia",
    "Grants": "Grants",
    "Models": "Models",
    "Websites": "Websites",
    "Reports/Analyses": "Reports/Analyses",
    "Posters/Conference Contribution": "Posters",
    "Book/Chapter/Collection": "Books/Collections",
    "Software Source Code": "Software Source Code",
    "Institution/Researcher Profiles": "Inst./Researcher Profiles",
}

# Count TRUE values
col_idx = {h: headers.index(h) for h in bool_cols}
counts = {}
for h in bool_cols:
    vals = [ws.cell(row=r, column=col_idx[h] + 1).value for r in range(2, ws.max_row + 1)]
    counts[h] = sum(1 for v in vals if v is True)

# Build ordered list with group colors
group_colors = {
    "Technical Features": "#4C72B0",
    "Metadata": "#DD8452",
    "Infrastructure": "#55A868",
    "Resource Types": "#C44E52",
}

feature_to_group = {c: g for g, cols in groups.items() for c in cols}

# Sort all features by count descending
sorted_features = sorted(bool_cols, key=lambda f: counts[f], reverse=True)

features = sorted_features
colors = [group_colors[feature_to_group[f]] for f in features]
pcts = [counts[f] / n_platforms * 100 for f in features]
labels = [short_labels[f] for f in features]

# Plot
fig, ax = plt.subplots(figsize=(10, 8))
y = np.arange(len(features))
bars = ax.barh(y, pcts, color=colors, edgecolor="white", linewidth=0.5)

# Add count labels
for bar, cnt in zip(bars, [counts[f] for f in features]):
    ax.text(bar.get_width() + 0.8, bar.get_y() + bar.get_height() / 2,
            f"{cnt}/{n_platforms}", va="center", ha="left", fontsize=8.5, color="#333333")

ax.set_yticks(y)
ax.set_yticklabels(labels, fontsize=9.5)
ax.invert_yaxis()
ax.set_xlabel("Percentage of platforms (%)", fontsize=11)
ax.set_title(f"Feature Prevalence Across NIH IC Data Ecosystem Platforms (n={n_platforms})",
             fontsize=12, fontweight="bold", pad=12)
ax.set_xlim(0, 110)
ax.axvline(50, color="gray", linestyle="--", linewidth=0.8, alpha=0.6)
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)


# Legend
legend_patches = [mpatches.Patch(color=c, label=g) for g, c in group_colors.items()]
ax.legend(handles=legend_patches, loc="lower right", fontsize=9, framealpha=0.8)

plt.tight_layout()
plt.savefig("analysis/feature_prevalence.png", dpi=300, bbox_inches="tight")
plt.savefig("analysis/feature_prevalence.pdf", bbox_inches="tight")
print("Saved: analysis/feature_prevalence.png and .pdf")
